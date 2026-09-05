#!/usr/bin/env python3
"""Read the provider's Evidence and Data Workbook and produce the analysis report.

    python3 _engine/analyze.py <slug>

Every number and every chart here comes from what the provider actually entered.
Where a sheet is empty this says so, in the report, in plain words, with what to
enter and where. It never fills a gap with a plausible-looking figure: a gap you
can explain survives a survey, and a fabricated record ends an agency.
"""
import json
import os
import sys
from collections import Counter, defaultdict
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import resolve as R
import chart_kit as CH
from build_provider import Sink, sub

SKILL = R.SKILL
QUARTERS = ["Q1", "Q2", "Q3", "Q4"]


# ------------------------------------------------------------------ reading
def read_sheet(wb, name):
    """Return a list of dicts. Header row is the first row whose cells are mostly text."""
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    hdr_row = None
    for r in range(1, min(5, ws.max_row) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        filled = [v for v in vals if v not in (None, "")]
        if len(filled) >= 3:
            hdr_row = r
            headers = [str(v).strip() if v is not None else "" for v in vals]
            break
    if hdr_row is None:
        return []
    out = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        row = {}
        any_val = False
        for c, h in enumerate(headers, 1):
            if not h:
                continue
            v = ws.cell(r, c).value
            row[h] = v
            if v not in (None, ""):
                any_val = True
        if any_val:
            out.append(row)
    return out


def num(v):
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace("%", "").replace("$", "").replace(",", "").strip())
    except Exception:
        return None


def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not v:
        return None
    for f in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(v).strip(), f).date()
        except Exception:
            pass
    return None


def yes(v):
    return str(v).strip().lower() in ("y", "yes", "true", "1")


def quarter_of(d, start):
    if not d or not start:
        return None
    m = (d.year - start.year) * 12 + (d.month - start.month)
    return m // 3 if 0 <= m < 12 else None


# ------------------------------------------------------------------ helpers
def no_data(s, what, sheet, why):
    s.notice(f"NO DATA ENTERED YET — {what}")
    s.md(f"Nothing has been entered on sheet **{sheet}** of the Evidence and Data Workbook. "
         f"{why}")
    s.md("**This is not a formatting problem and it cannot be written around.** A surveyor "
         "asks for this by name. Enter the real records as they happen and re-run "
         "`python3 _engine/analyze.py <slug>`.")


def section(s, title, area, plain):
    s.heading(title, 1)
    s.rule()
    s.md(f"**CARF area {area}.** {plain}")


def stat_tiles(s, tiles):
    """A hero row. Numbers, not a chart - a chart of four numbers is noise."""
    s.table([[t[0] for t in tiles], [t[1] for t in tiles]])


# ------------------------------------------------------------------ sections
def sec_completeness(s, sheets, tk):
    section(s, "1. Is there enough evidence yet?", "1.M / 1.N",
            "Before any analysis is credible, this says how much real data exists. A surveyor "
            "forms the same judgement in the first ten minutes.")
    rows = [["Sheet", "Rows entered", "What it proves", "Enough for survey?"]]
    expect = [
        ("03", "Drills", "Emergency preparedness actually tested", "4+ per shift per year, one unannounced"),
        ("04", "Incidents", "Incidents captured, reviewed and analysed", "Every reportable event"),
        ("05", "Grievances", "Complaints heard and answered on time", "Every concern, informal included"),
        ("06", "Record Reviews", "Documentation checked before a payer does", "4 quarters"),
        ("07", "Access and Referrals", "People get in, and how fast", "Every referral"),
        ("08", "Satisfaction — persons served", "The people served were asked", "1 annual cycle"),
        ("09", "Satisfaction — stakeholders", "Stakeholders were asked", "1 annual cycle"),
        ("10", "Outcomes — persons served", "Services changed lives measurably", "Every person served"),
        ("11", "Personnel", "Staff qualified, trained, supervised", "Every employee"),
        ("12", "Financial", "Budget monitored, controls working", "12 months"),
        ("13", "Accessibility", "Barriers found and removed", "All 8 areas"),
        ("14", "PI Projects", "Improvement actually happened", "1+ complete with before/after"),
        ("15", "Measure Results", "Performance measured against target", "4 quarters"),
    ]
    filled = 0
    for pref, name, proves, enough in expect:
        n = len(sheets.get(pref, []))
        if n:
            filled += 1
        rows.append([f"{pref} {name}", str(n), proves,
                     enough if not n else ("Yes" if n >= 4 else "Partial")])
    s.table(rows)
    pct = round(100 * filled / len(expect))
    s.md(f"**{filled} of {len(expect)} evidence sheets have any data in them ({pct}%).** "
         "Empty sheets are not a formatting gap; they are the work that has not happened yet.")


def sec_checklist(s, cal_rows, outdir, tk):
    section(s, "2. Where the agency stands against every required item", "All",
            "Every dated obligation in the packet, counted by status. This is the honest "
            "answer to \"are we ready?\".")
    counts = Counter(r["status"] for r in cal_rows)
    p = CH.status_bar(os.path.join(outdir, "charts", "checklist_status.png"),
                      {"Upcoming": counts.get("Upcoming", 0),
                       "Due within 30 days": counts.get("Due within 30 days", 0),
                       "OVERDUE": counts.get("OVERDUE", 0)},
                      "Required items by status",
                      "Every drill, review, survey, plan and report the packet tracks")
    s.image(p)
    overdue = [r for r in cal_rows if r["status"] == "OVERDUE"]
    if overdue:
        s.md(f"**{len(overdue)} item(s) are overdue.** They are listed below, worst first. "
             "Every one of these is something a surveyor can ask for by name.")
        rows = [["Days late", "Due", "Item", "Which one", "Owner", "Evidence expected"]]
        for r in sorted(overdue, key=lambda x: -x["days"])[:40]:
            rows.append([str(r["days"]), str(r["due"]), r["item"], r["instance"],
                         r["owner"], r["evidence"]])
        s.table(rows)
        if len(overdue) > 40:
            s.md(f"...and {len(overdue) - 40} more. The full list is sheet "
                 "**01 MASTER CHECK-OFF LIST**, filtered to Status = OVERDUE.")
    else:
        s.md("**Nothing is overdue.** Keep it that way — the calendar is the easiest part of "
             "an accreditation to let slip and the easiest for a surveyor to catch.")
    by_dom = defaultdict(lambda: [0, 0])
    for r in cal_rows:
        by_dom[r["domain"]][0] += 1
        if r["status"] == "OVERDUE":
            by_dom[r["domain"]][1] += 1
    if any(v[1] for v in by_dom.values()):
        p = CH.ranked_bar(os.path.join(outdir, "charts", "overdue_by_domain.png"),
                          [k.title() for k in by_dom], [v[1] for v in by_dom.values()],
                          "Overdue items by section", "Where the agency is behind",
                          xlabel="Overdue items", color=CH.STATUS["critical"])
        s.image(p)


def sec_measures(s, sheets, outdir, tk):
    section(s, "3. Performance measures against target", "1.M",
            "Every measure in the Performance Measurement Plan, with its real quarterly "
            "result. A measure with no target cannot be passed or failed, and a target you "
            "invented and missed is worse than a modest one you met.")
    rows = sheets.get("15", [])
    if not rows:
        return no_data(s, "no measure results", "15 Measure Results",
                       "Without this there is no performance analysis, no dashboard, and "
                       "nothing to show for CARF areas 1.M and 1.N.")
    table = [["ID", "Measure", "Target", "Q1", "Q2", "Q3", "Q4", "Year", "Prior year", "Met?"]]
    charted = 0
    for r in rows:
        vals = [num(r.get(q)) for q in QUARTERS]
        target = num(r.get("Target"))
        year = num(r.get("Year"))
        if all(v is None for v in vals) and year is None:
            met = "no data"
        elif target is None:
            met = "no target set"
        else:
            latest = year if year is not None else next((v for v in reversed(vals) if v is not None), None)
            met = "—" if latest is None else ("Yes" if latest >= target else "No")
        table.append([str(r.get("Measure ID", "")), str(r.get("Measure", "")),
                      "" if target is None else f"{target:g}",
                      *[("" if v is None else f"{v:g}") for v in vals],
                      "" if year is None else f"{year:g}",
                      str(r.get("Prior year") or ""), met])
        if any(v is not None for v in vals) and charted < 8:
            name = str(r.get("Measure", ""))[:46]
            p = CH.trend(os.path.join(outdir, "charts",
                                      f"m_{str(r.get('Measure ID','m'))}.png"),
                         QUARTERS, {name: vals},
                         f"{r.get('Measure ID','')} — {name}",
                         f"{r.get('Domain','')} · unit: {r.get('Unit','')}",
                         target=target, ylabel=str(r.get("Unit") or ""))
            r["_chart"] = p
            charted += 1
    s.table(table)
    s.md("")
    for r in rows:
        if r.get("_chart"):
            s.image(r["_chart"])
    missed = [r for r in table[1:] if r[-1] == "No"]
    if missed:
        s.md(f"**{len(missed)} measure(s) missed target.** Each one needs a probable cause and "
             "a corrective action written into section 12 of this report. A missed target that "
             "is analysed is a strength at survey. A missed target that is silently omitted is "
             "the finding.")


def sec_drills(s, sheets, provider, outdir, start):
    section(s, "4. Emergency drills — coverage, not just count", "1.H",
            "Fire and evacuation drills quarterly on EVERY shift that delivers service, every "
            "emergency type in the plan at least annually, and at least one unannounced.")
    rows = sheets.get("03", [])
    if not rows:
        return no_data(s, "no drills logged", "03 Drills",
                       "Drill records are among the first things a surveyor asks for, and the "
                       "most common place an otherwise strong agency is found short.")
    shifts = provider.get("shifts") or ["All service hours"]
    by_shift_q = {sh: [0, 0, 0, 0] for sh in shifts}
    types = Counter()
    unannounced = 0
    no_critique = 0
    for r in rows:
        d = as_date(r.get("Date"))
        q = quarter_of(d, start)
        sh = str(r.get("Shift") or shifts[0]).strip()
        if sh in by_shift_q and q is not None:
            by_shift_q[sh][q] += 1
        types[str(r.get("Drill type") or "unspecified").strip()] += 1
        if "unannounce" in str(r.get("Announced or unannounced") or "").lower():
            unannounced += 1
        if not str(r.get("What went wrong") or "").strip():
            no_critique += 1
    stat_tiles(s, [("Drills logged", str(len(rows))),
                   ("Unannounced", str(unannounced)),
                   ("Distinct types drilled", str(len(types))),
                   ("Missing a critique", str(no_critique))])
    p = CH.grouped_bar(os.path.join(outdir, "charts", "drills_by_shift.png"),
                       QUARTERS, {sh: by_shift_q[sh] for sh in list(by_shift_q)[:3]},
                       "Drills completed by quarter and shift",
                       "Quarterly fire/evacuation drills are required on every shift that delivers service",
                       ylabel="Drills", target=1)
    s.image(p)
    p = CH.ranked_bar(os.path.join(outdir, "charts", "drill_types.png"),
                      list(types), list(types.values()),
                      "Drills by emergency type", "Every type in the Health and Safety Plan needs at least one",
                      xlabel="Drills run")
    s.image(p)
    gaps = []
    for sh, qs in by_shift_q.items():
        empty = [QUARTERS[i] for i, v in enumerate(qs) if v == 0]
        if empty:
            gaps.append(f"**{sh}**: no drill logged in {', '.join(empty)}")
    if unannounced == 0:
        gaps.append("**No unannounced drill** has been logged. A drill everyone knew about is "
                    "not a test, and surveyors look for this specifically.")
    if no_critique:
        gaps.append(f"**{no_critique} drill(s) have no 'what went wrong' entry.** A drill with "
                    "nothing to fix reads as a drill nobody watched.")
    if gaps:
        s.notice("DRILL GAPS TO CLOSE")
        for g in gaps:
            s.md(f"- {g}")
    else:
        s.md("**Drill coverage is complete** for every shift and quarter with data entered.")


def sec_incidents(s, sheets, outdir, start):
    section(s, "5. Critical incidents — trend and follow-through", "1.H",
            "Incidents aggregated at least quarterly by type, location, time and staff, with a "
            "written conclusion even when the conclusion is that no trend exists.")
    rows = sheets.get("04", [])
    if not rows:
        return no_data(s, "no incidents logged", "04 Incidents",
                       "An empty incident log is read as an unused log, not as a safe year. If "
                       "genuinely nothing happened, that fact still has to be recorded and analysed.")
    types = Counter(str(r.get("Type") or "unspecified").strip() for r in rows)
    by_q = [0, 0, 0, 0]
    late_review, no_ext = 0, 0
    for r in rows:
        q = quarter_of(as_date(r.get("Date")), start)
        if q is not None:
            by_q[q] += 1
        d, rev = as_date(r.get("Date")), as_date(r.get("Reviewed within 5 business days (date)"))
        if d and (rev is None or (rev - d).days > 7):
            late_review += 1
        if str(r.get("External report required?") or "").strip().lower().startswith("y") \
                and not as_date(r.get("External report made (date)")):
            no_ext += 1
    stat_tiles(s, [("Incidents", str(len(rows))), ("Types seen", str(len(types))),
                   ("Not reviewed within 5 days", str(late_review)),
                   ("External report missing", str(no_ext))])
    s.image(CH.ranked_bar(os.path.join(outdir, "charts", "incident_types.png"),
                          list(types), list(types.values()),
                          "Incidents by type", "Ranked — the top bar is where prevention effort belongs",
                          xlabel="Incidents", color=CH.STATUS["serious"]))
    s.image(CH.trend(os.path.join(outdir, "charts", "incident_trend.png"), QUARTERS,
                     {"Incidents": by_q}, "Incidents by quarter",
                     "Direction matters more than the absolute number", ylabel="Incidents",
                     lower_is_better=True))
    if late_review or no_ext:
        s.notice("FOLLOW-THROUGH GAPS")
        if late_review:
            s.md(f"- **{late_review} incident(s)** have no supervisor review within five business days.")
        if no_ext:
            s.md(f"- **{no_ext} incident(s)** needed an external report with no date recorded. "
                 "A missed mandatory report is a regulatory matter before it is an accreditation one.")


def sec_grievances(s, sheets, outdir, start):
    section(s, "6. Grievances — heard, answered, on time", "1.K",
            "Written acknowledgment within three business days, a written decision within thirty "
            "calendar days, and a quarterly trend analysis.")
    rows = sheets.get("05", [])
    if not rows:
        return no_data(s, "no grievances logged", "05 Grievances",
                       "An empty grievance log is one of the strongest negative signals in a "
                       "survey. It usually means informal concerns are not being captured, not "
                       "that nobody ever complained.")
    cats = Counter(str(r.get("Category") or "uncategorised").strip() for r in rows)
    on_time = sum(1 for r in rows if yes(r.get("Within 30 days?")))
    ack_time = sum(1 for r in rows if yes(r.get("Within 3 business days?")))
    n = len(rows)
    stat_tiles(s, [("Grievances", str(n)),
                   ("Acknowledged in 3 days", f"{round(100*ack_time/n)}%"),
                   ("Decided within 30 days", f"{round(100*on_time/n)}%"),
                   ("Categories seen", str(len(cats)))])
    s.image(CH.ranked_bar(os.path.join(outdir, "charts", "grievance_cats.png"),
                          list(cats), list(cats.values()), "Grievances by category",
                          "The tallest bar is the systemic issue to fix, not the person to blame",
                          xlabel="Grievances"))
    if on_time < n:
        s.notice(f"{n - on_time} grievance(s) were not decided within thirty days.")
        s.md("Timeframes in the grievance procedure are a promise to the people served. Missing "
             "them is a rights finding, not an administrative one.")


def sec_records(s, sheets, outdir, start):
    section(s, "7. Quality record review", "2.H",
            "Records reviewed quarterly against a written tool by someone who did not write "
            "the documentation, with every finding tracked to a verified correction.")
    rows = sheets.get("06", [])
    if not rows:
        return no_data(s, "no record reviews logged", "06 Record Reviews",
                       "Four quarters of record review is the evidence that the agency polices "
                       "its own documentation. Without it, the payer or the surveyor does it first.")
    by_q = defaultdict(list)
    for r in rows:
        q = str(r.get("Quarter") or "").strip() or (QUARTERS[quarter_of(as_date(r.get("Review date")), start)]
                                                    if quarter_of(as_date(r.get("Review date")), start) is not None else "")
        sc = num(r.get("Score %"))
        if q and sc is not None:
            by_q[q].append(sc)
    means = [round(sum(by_q[q]) / len(by_q[q]), 1) if by_q.get(q) else None for q in QUARTERS]
    defects = Counter()
    for r in rows:
        for col, lab in [("Goal linkage OK?", "Goal linkage"), ("Notes timely?", "Note timeliness"),
                         ("Person's own words?", "Person's own words"),
                         ("Billing matches documentation?", "Billing matches documentation")]:
            v = str(r.get(col) or "").strip().lower()
            if v.startswith("n") and v != "n/a":
                defects[lab] += 1
        if str(r.get("Cloned text?") or "").strip().lower().startswith("y"):
            defects["Cloned text"] += 1
    verified = sum(1 for r in rows if as_date(r.get("Date verified")))
    with_find = sum(1 for r in rows if str(r.get("Findings") or "").strip())
    stat_tiles(s, [("Records reviewed", str(len(rows))),
                   ("With findings", str(with_find)),
                   ("Corrections verified", str(verified)),
                   ("Quarters covered", str(sum(1 for m in means if m is not None)))])
    s.image(CH.trend(os.path.join(outdir, "charts", "record_scores.png"), QUARTERS,
                     {"Mean review score": means}, "Record review score by quarter",
                     "Rising is good; flat with findings still open is not", ylabel="%", target=90))
    if defects:
        s.image(CH.ranked_bar(os.path.join(outdir, "charts", "record_defects.png"),
                              list(defects), list(defects.values()),
                              "Most common documentation defects",
                              "Fix the top bar with training, not with more review",
                              xlabel="Records affected", color=CH.SERIES[1]))
    if with_find and verified < with_find:
        s.notice(f"{with_find - verified} finding(s) have no verification date.")
        s.md("A correction is not closed on a promise. Somebody has to re-examine the "
             "documentation and date it. This is the single most common reason a quality "
             "system looks good on paper and fails at survey.")


def sec_access(s, sheets, outdir, start, tk):
    section(s, "8. Access to services", "2.B",
            "How long people wait, how many engage, and what happens to the people who are "
            "turned away.")
    rows = sheets.get("07", [])
    if not rows:
        return no_data(s, "no referral data", "07 Access and Referrals",
                       "Access measures (A1-A4) cannot be reported without it, and 'how long "
                       "from referral to first contact?' is asked in almost every survey.")
    by_q = defaultdict(list)
    engaged = 0
    declined = []
    for r in rows:
        q = quarter_of(as_date(r.get("Date received")), start)
        d = num(r.get("Days to first contact"))
        if q is not None and d is not None:
            by_q[q].append(d)
        if yes(r.get("Engaged (2+ contacts)?")):
            engaged += 1
        rsn = str(r.get("Not accepted — reason") or "").strip()
        if rsn:
            declined.append(rsn)
    means = [round(sum(by_q[i]) / len(by_q[i]), 1) if by_q.get(i) else None for i in range(4)]
    n = len(rows)
    stat_tiles(s, [("Referrals", str(n)),
                   ("Engaged (2+ contacts)", f"{round(100*engaged/n)}%"),
                   ("Not accepted", str(len(declined))),
                   ("Target days to contact", tk["ACCESS_DAYS"])])
    s.image(CH.trend(os.path.join(outdir, "charts", "access_days.png"), QUARTERS,
                     {"Mean days referral to first contact": means},
                     "Days from referral to first contact",
                     "Against the agency's own stated access standard",
                     target=num(tk["ACCESS_DAYS"]), ylabel="days", lower_is_better=True))
    if declined:
        c = Counter(declined)
        s.image(CH.ranked_bar(os.path.join(outdir, "charts", "declined.png"),
                              list(c), list(c.values()), "Reasons people were not accepted",
                              "Check every reason here is about capability, never a protected characteristic",
                              xlabel="People"))


def sec_satisfaction(s, sheets, outdir):
    section(s, "9. Satisfaction — persons served and stakeholders", "1.D",
            "Both groups must be asked, analysed, and told what changed as a result.")
    ps = sheets.get("08", [])
    st = sheets.get("09", [])
    if not ps and not st:
        return no_data(s, "no satisfaction data", "08 / 09 Satisfaction",
                       "Input from persons served AND from other stakeholders are two separate "
                       "requirements. Both are needed.")
    for label, rows, fname in [("Persons served", ps, "sat_ps.png"),
                               ("Other stakeholders", st, "sat_stake.png")]:
        if not rows:
            s.md(f"**{label}: no responses entered.** This half of the requirement is unmet.")
            continue
        qcols = [k for k in rows[0] if k.startswith("Q")]
        labels, means = [], []
        for k in qcols:
            vals = [num(r.get(k)) for r in rows]
            vals = [v for v in vals if v is not None]
            if vals:
                labels.append(k)
                means.append(round(sum(vals) / len(vals), 2))
        if labels:
            s.image(CH.ranked_bar(os.path.join(outdir, "charts", fname), labels, means,
                                  f"{label} — mean score by question",
                                  f"{len(rows)} response(s). Scale 1-5. The lowest bar is where to act.",
                                  xlabel="Mean score (1-5)", color=CH.SERIES[2], top=20))
            low = min(zip(means, labels))
            s.md(f"**Lowest-scoring question: {low[1]} at {low[0]}.** That is the improvement "
                 "action this survey is asking for.")


def sec_outcomes(s, sheets, outdir):
    section(s, "10. Outcomes for the people served", "1.M / 3.CI",
            "Whether services actually changed lives — goals reached, housing, work, and "
            "whether people's unpaid support networks grew.")
    rows = sheets.get("10", [])
    if not rows:
        return no_data(s, "no outcome data", "10 Outcomes — persons served",
                       "This sheet produces almost every effectiveness measure. Without it the "
                       "agency can describe what it did but not what changed.")
    n = len(rows)
    goals_set = sum(num(r.get("Goals set")) or 0 for r in rows)
    goals_ach = sum(num(r.get("Goals achieved")) or 0 for r in rows)
    with_goal = sum(1 for r in rows if (num(r.get("Goals achieved")) or 0) > 0)
    adm = [num(r.get("Natural supports at admission")) for r in rows]
    lastr = [num(r.get("Natural supports at last review")) for r in rows]
    pairs = [(a, b) for a, b in zip(adm, lastr) if a is not None and b is not None]
    grew = sum(1 for a, b in pairs if b > a)
    stat_tiles(s, [("People served", str(n)),
                   ("Achieved 1+ goal", f"{round(100*with_goal/n)}%"),
                   ("Goals achieved / set", f"{goals_ach:g} / {goals_set:g}"),
                   ("Natural supports grew", f"{round(100*grew/len(pairs))}%" if pairs else "no data")])
    dom = [("Housing", "Housing goal? Achieved?"),
           ("Employment / education", "Employment/education goal? Achieved?"),
           ("Benefits", "Benefits goal? Achieved?")]
    labs, vals = [], []
    for lab, col in dom:
        have = [r for r in rows if str(r.get(col) or "").strip()]
        ach = [r for r in have if yes(str(r.get(col)).split("/")[-1]) or "achieved" in str(r.get(col)).lower()]
        if have:
            labs.append(lab)
            vals.append(round(100 * len(ach) / len(have)))
    if labs:
        s.image(CH.ranked_bar(os.path.join(outdir, "charts", "outcome_domains.png"), labs, vals,
                              "Goal achievement by life domain",
                              "Of the people who had a goal in that domain", xlabel="% achieved",
                              color=CH.SERIES[2]))
    if pairs:
        s.image(CH.paired_change(os.path.join(outdir, "charts", "supports.png"),
                                 ["Mean unpaid supports"],
                                 [round(sum(a for a, _ in pairs) / len(pairs), 1)],
                                 [round(sum(b for _, b in pairs) / len(pairs), 1)],
                                 "Unpaid natural supports, admission to latest review",
                                 "Paid support ends; relationships do not. This is the measure that "
                                 "shows community integration actually happened.",
                                 ylabel="Mean number of unpaid supports"))


def sec_personnel(s, sheets, outdir):
    section(s, "11. Workforce compliance", "1.I / 3.PEER",
            "Credentials verified and current, orientation and competency complete, supervision "
            "happening, appraisals done.")
    rows = sheets.get("11", [])
    if not rows:
        return no_data(s, "no personnel data", "11 Personnel",
                       "A surveyor pulls three personnel files at random. This sheet tells you "
                       "in advance which three would hurt.")
    active = [r for r in rows if not as_date(r.get("Separation date"))]
    n = len(active) or 1
    today = date.today()

    def pct(f):
        return round(100 * sum(1 for r in active if f(r)) / n)

    checks = [
        ("Credential on file and unexpired",
         lambda r: (as_date(r.get("Credential expires")) or date(1900, 1, 1)) >= today),
        ("Orientation completed", lambda r: as_date(r.get("Orientation completed")) is not None),
        ("Initial competency done", lambda r: as_date(r.get("Competency — initial")) is not None),
        ("Annual competency current",
         lambda r: (as_date(r.get("Competency — last annual")) or date(1900, 1, 1)) > _yr_ago(today)),
        ("Direct observation recorded",
         lambda r: as_date(r.get("Direct observation (last)")) is not None),
        ("Appraisal within 12 months",
         lambda r: (as_date(r.get("Appraisal — last")) or date(1900, 1, 1)) > _yr_ago(today)),
        ("Annual training complete", lambda r: yes(r.get("Annual training complete?"))),
        ("Disclosure plan on file", lambda r: as_date(r.get("Disclosure plan (date)")) is not None),
        ("Wellness plan on file", lambda r: as_date(r.get("Wellness plan (date)")) is not None),
        ("First aid / CPR current",
         lambda r: (as_date(r.get("First aid/CPR expires")) or date(1900, 1, 1)) >= today),
    ]
    labs = [c[0] for c in checks]
    vals = [pct(c[1]) for c in checks]
    s.image(CH.ranked_bar(os.path.join(outdir, "charts", "personnel.png"), labs, vals,
                          "Workforce compliance",
                          f"{len(active)} active staff. Anything under 100% is a file a surveyor could pull.",
                          xlabel="% of active staff", top=12))
    weak = [(l, v) for l, v in zip(labs, vals) if v < 100]
    if weak:
        s.notice("FILES THAT WOULD NOT SURVIVE A PULL")
        rows2 = [["Requirement", "% complete", "Staff missing it"]]
        for lab, v in sorted(weak, key=lambda x: x[1]):
            f = dict(checks)[lab]
            missing = [str(r.get("Name") or r.get("Role") or "?") for r in active if not f(r)]
            rows2.append([lab, f"{v}%", ", ".join(missing[:10]) or "—"])
        s.table(rows2)


def _yr_ago(d):
    try:
        return d.replace(year=d.year - 1)
    except ValueError:
        return d.replace(year=d.year - 1, day=28)


def sec_accessibility(s, sheets, outdir):
    section(s, "12. Accessibility progress", "1.L",
            "Barriers found across all eight areas, and at least one removed or reduced this year.")
    rows = sheets.get("13", [])
    if not rows:
        return no_data(s, "no accessibility data", "13 Accessibility",
                       "An accessibility plan with no identified barriers reads as a plan nobody "
                       "worked. Every area needs a row, even if the row says 'none identified'.")
    areas = ["Architectural", "Environmental", "Attitudinal", "Financial", "Employment",
             "Communication", "Transportation", "Digital"]
    found = Counter(str(r.get("Area (of the 8)") or "").strip().title() for r in rows)
    removed = sum(1 for r in rows if as_date(r.get("Date removed/reduced")))
    missing = [a for a in areas if not any(a.lower() in k.lower() for k in found if k)]
    stat_tiles(s, [("Barriers logged", str(len(rows))), ("Removed or reduced", str(removed)),
                   ("Areas covered", f"{8 - len(missing)} of 8"),
                   ("Areas not examined", str(len(missing)))])
    s.image(CH.ranked_bar(os.path.join(outdir, "charts", "access_barriers.png"),
                          [a for a in areas], [found.get(a, 0) for a in areas],
                          "Barriers identified by accessibility area",
                          "An area with zero is either genuinely clear or was never examined — say which",
                          xlabel="Barriers logged", color=CH.SERIES[1]))
    if missing:
        s.notice("AREAS WITH NO ENTRY: " + ", ".join(missing))
        s.md("Add a row for each, even if it reads \"None identified\" — and say how you looked.")
    if removed == 0:
        s.notice("No barrier has been recorded as removed or reduced this year.")


def sec_pi(s, sheets):
    section(s, "13. Performance improvement projects", "1.N",
            "At least one completed project with before-and-after numbers is what turns "
            "\"we improve\" from a claim into evidence.")
    rows = sheets.get("14", [])
    if not rows:
        return no_data(s, "no improvement projects", "14 PI Projects",
                       "This is the evidence that measurement led to change. Without it the "
                       "quality system stops at counting.")
    t = [["#", "Problem", "Aim", "Baseline", "End", "Decision", "What changed permanently"]]
    for r in rows:
        t.append([str(r.get("Project #") or ""), str(r.get("Problem (with a number)") or ""),
                  str(r.get("Aim (from X to Y by when)") or ""), str(r.get("Baseline") or ""),
                  str(r.get("End result") or ""), str(r.get("Decision (adopt/adapt/abandon)") or ""),
                  str(r.get("What changed permanently") or "")])
    s.table(t)
    done = [r for r in rows if num(r.get("Baseline")) is not None and num(r.get("End result")) is not None]
    if not done:
        s.notice("No project has both a baseline and an end result recorded.")
    else:
        s.md(f"**{len(done)} project(s) have complete before-and-after data.** Those are the ones "
             "to walk a surveyor through.")


def sec_conclusions(s, tk):
    section(s, "14. Written conclusions, actions and distribution", "1.N",
            "The analysis above is data. This section is where leadership says what it MEANS "
            "and what will change — and it has to be written by a person, not generated.")
    s.notice("THIS SECTION IS DELIBERATELY BLANK. IT IS YOURS TO WRITE.")
    s.md("A generated conclusion would be a fabricated one. The numbers above are real; the "
         "judgement about them has to come from the people who run the agency, and a surveyor "
         "will ask the CEO to explain it in their own words.")
    for h, prompt in [
        ("What is working, and how we know", "Name the measure and the evidence."),
        ("What is not working, and the probable cause", "Be specific. \"Staffing\" is not a cause."),
        ("Where we met target and will now raise it", "Meeting a target is a result, not an endpoint."),
        ("What surprised us", "Surveyors like this question. It shows the analysis was actually read."),
    ]:
        s.heading(h, 2)
        s.md(f"*{prompt}*")
        s.md("_______________________________________________________________________________")
        s.md("_______________________________________________________________________________")
    s.heading("Improvement actions for the coming year", 2)
    s.table([["#", "Action", "Which finding it addresses", "Owner", "Target date",
              "How we will know it worked"],
             ["1", "", "", "", "", ""], ["2", "", "", "", "", ""], ["3", "", "", "", "", ""]])
    s.heading("Distribution — the most-missed requirement in the manual", 2)
    s.md("Analysing performance is half of area 1.N. Telling people the results is the other "
         "half, and it is where most agencies are found short. Complete this within 90 days.")
    s.table([["Audience", "Format", "Date shared", "How it is documented"],
             ["Personnel", "Staff meeting + written summary", "", "Minutes, sign-in sheet"],
             ["Persons served", "Plain-language one-page summary", "", "Posted + handed out; log"],
             [sub("{{GOVERNING_BODY}}", {"GOVERNING_BODY": tk["GOVERNING_BODY"]}), "Full report", "", "Minutes"],
             ["Referral sources and payers", "Summary", "", "Email record"]])
    s.md("")
    s.md("Prepared by: ________________________  Signature: ________________________  Date: __________")
    s.md("Reviewed by: ________________________  Signature: ________________________  Date: __________")


# ------------------------------------------------------------------ main
def main():
    if len(sys.argv) < 2:
        sys.exit("usage: analyze.py <slug>")
    slug = sys.argv[1]
    provider = R.load_provider(slug)
    provider.setdefault("slug", slug)
    res = R.Resolver(provider)
    res.gate()
    tk = res.tokens()
    tk["SLUG"] = slug
    verified = bool(provider.get("manual_verified"))

    from openpyxl import load_workbook
    import calendar_engine as CE

    data_dir = os.path.join(SKILL, "providers", slug, "data")
    wb_path = os.path.join(data_dir, "Evidence_and_Data_Workbook.xlsx")
    if not os.path.exists(wb_path):
        sys.exit(f"No workbook at {wb_path}\nRun: python3 _engine/build_provider.py {slug}")
    wb = load_workbook(wb_path, data_only=True)
    sheets = {n[:2]: read_sheet(wb, n) for n in wb.sheetnames}
    sheets.update({n: read_sheet(wb, n) for n in wb.sheetnames})
    cal_rows, start, _ = CE.expand(provider)

    outdir = os.path.join(SKILL, "providers", slug, "output")
    os.makedirs(os.path.join(outdir, "charts"), exist_ok=True)
    # Count only ENTERED data: the event logs 03-14, plus any measure row that
    # actually carries a quarterly figure. The seeded measure skeleton and the
    # evidence register are scaffolding, not evidence, and must not inflate this.
    total_rows = sum(len(v) for k, v in sheets.items()
                     if len(k) == 2 and k.isdigit() and "03" <= k <= "14")
    total_rows += sum(1 for r in sheets.get("15", [])
                      if any(num(r.get(q)) is not None for q in QUARTERS + ["Year"]))

    for kind, ext in (("docx", ".docx"), ("pdf", ".pdf")):
        s = Sink(kind)
        s.title_page(tk["AGENCY"], "Performance Analysis and Evidence Report",
                     "Built from the agency's own entered data — nothing here is estimated",
                     [f"Program: {tk['PROGRAM_NAME']}",
                      f"Data rows analysed: {total_rows}",
                      f"Generated {tk['TODAY']}"],
                     warning=None if verified else
                     "DRAFT - the CARF section map behind this report has NOT been checked against the purchased manual.")
        s.heading("How to read this report", 1)
        s.md("Every figure and every chart below was computed from the Evidence and Data "
             "Workbook. Where a sheet is empty, this report says so rather than estimating. "
             "**Re-run it whenever you enter more data** — it is designed to be run monthly, "
             "not once.")
        s.md("`python3 _engine/analyze.py " + slug + "`")
        s.pb()
        sec_completeness(s, sheets, tk); s.pb()
        sec_checklist(s, cal_rows, outdir, tk); s.pb()
        sec_measures(s, sheets, outdir, tk); s.pb()
        sec_drills(s, sheets, provider, outdir, start); s.pb()
        sec_incidents(s, sheets, outdir, start); s.pb()
        sec_grievances(s, sheets, outdir, start); s.pb()
        sec_records(s, sheets, outdir, start); s.pb()
        sec_access(s, sheets, outdir, start, tk); s.pb()
        sec_satisfaction(s, sheets, outdir); s.pb()
        sec_outcomes(s, sheets, outdir); s.pb()
        sec_personnel(s, sheets, outdir); s.pb()
        sec_accessibility(s, sheets, outdir); s.pb()
        sec_pi(s, sheets); s.pb()
        sec_conclusions(s, tk)
        foot = (f"{tk['AGENCY_SHORT']} — Performance Analysis — "
                f"{'' if verified else 'DRAFT, section map unverified — '}{tk['TODAY']}")
        path = os.path.join(outdir, "07_Performance_Analysis_Report" + ext)
        s.save(path, foot)
        print(f"  wrote 07_Performance_Analysis_Report{ext}")
    print(f"\nAnalysed {total_rows} data rows from {wb_path}")
    print(f"Report: {outdir}/07_Performance_Analysis_Report.pdf")


if __name__ == "__main__":
    main()
