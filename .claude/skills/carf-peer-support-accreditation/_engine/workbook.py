"""Build the Evidence and Data Workbook.

This is where the provider's REAL data lives. Nothing in it is pre-filled with
invented results: the logs ship empty with worked column headers, and analyze.py
reads back whatever the provider actually enters. Fabricating a drill date or a
satisfaction score would be the one thing that turns an accreditation packet into
a fraud, so the workbook is built to make entry easy and invention unnecessary.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NAVY = "1F3B63"
HEAD = PatternFill("solid", fgColor=NAVY)
SUB = PatternFill("solid", fgColor="E8EDF5")
RED = PatternFill("solid", fgColor="FBE3E3")
AMBER = PatternFill("solid", fgColor="FFF3D9")
GREEN = PatternFill("solid", fgColor="E4F3E6")
THIN = Border(*[Side(style="thin", color="C9D2E0")] * 4)


def _sheet(wb, title, headers, widths, notes=None, freeze="A2"):
    ws = wb.create_sheet(title[:31])
    r = 1
    if notes:
        ws.cell(1, 1, notes).font = Font(italic=True, size=9, color="55606E")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 4))
        ws.row_dimensions[1].height = 26
        ws.cell(1, 1).alignment = Alignment(wrap_text=True, vertical="center")
        r = 2
        freeze = f"A{r + 1}"
    ws.append([]) if r == 2 else None
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = HEAD
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A{r}:{get_column_letter(len(headers))}{r}"
    ws._hdr_row = r
    return ws


def _yesno(ws, col, first=3, last=600):
    dv = DataValidation(type="list", formula1='"Yes,No,N/A"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}{first}:{col}{last}")


def build(path, tk, provider, cal_rows, cal_start, cal_end):
    wb = Workbook()
    ws = wb.active
    ws.title = "00 How to use"

    intro = [
        (f"{tk['AGENCY']} — CARF Evidence and Data Workbook", 16, True, NAVY),
        (f"Built {tk['TODAY']}.  Target survey: {tk['SURVEY_TARGET']}.", 10, False, "55606E"),
        ("", 10, False, "000000"),
        ("This workbook is the spine of the whole accreditation. Everything a surveyor asks "
         "for is either a document in the packet or a number in here.", 11, True, "000000"),
        ("", 10, False, "000000"),
        ("HOW TO USE IT", 12, True, NAVY),
        ("1.  '01 MASTER CHECK-OFF LIST' has every required item with the date it is due. Work "
         "it weekly. Put a real date in 'Date completed' and say where the evidence is filed.", 10, False, "000000"),
        ("2.  '02 Compliance Calendar' is the same list arranged by month, so you can see what "
         "is due in March without scrolling.", 10, False, "000000"),
        ("3.  Sheets 03-14 are the LOGS. Enter real events as they happen — a drill the day you "
         "run it, an incident the day it occurs. Do not batch them at the end.", 10, False, "000000"),
        ("4.  '15 Measure Results' is what the charts and the annual analysis are built from. "
         "Enter the actual quarterly number for every measure.", 10, False, "000000"),
        ("5.  '16 Evidence Register' lists every document you hold, its date, and where it "
         "lives — this becomes your survey-day index.", 10, False, "000000"),
        ("6.  When there is data in here, run:   python3 _engine/analyze.py <slug>", 10, True, "000000"),
        ("     That produces the Performance Analysis Report with real charts, built only from "
         "what you entered.", 10, False, "000000"),
        ("", 10, False, "000000"),
        ("THE ONE RULE", 12, True, "A81C1C"),
        ("Never write a date for something that did not happen, and never enter a number you "
         "cannot trace to a source. A gap you can explain is survivable at survey. A fabricated "
         "record is fraud, and it is the one thing that ends an agency.", 10, True, "A81C1C"),
        ("An honest 'not done yet' in this workbook is exactly what a good consultant wants to "
         "see — it is the list of what to fix before the surveyor arrives.", 10, False, "000000"),
    ]
    for i, (text, size, bold, color) in enumerate(intro, 1):
        c = ws.cell(i, 1, text)
        c.font = Font(size=size, bold=bold, color=color)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 30 if size >= 12 else (14 if not text else 28)
    ws.column_dimensions["A"].width = 118

    # ---------- 01 master check-off list ----------
    dom_names = {d["key"]: d["name"] for d in __import__("json").load(
        open(__file__.replace("workbook.py", "content/meta/obligations.json")))["domains"]}
    ck = _sheet(wb, "01 MASTER CHECK-OFF LIST",
                ["Due date", "Status", "Days", "Section (domain)", "CARF area", "Item",
                 "Which one / who", "Period", "Owner", "Evidence that proves it",
                 "Done?", "Date completed", "Where the evidence is filed", "Notes"],
                [11, 20, 7, 26, 9, 52, 22, 14, 20, 46, 9, 13, 30, 30],
                notes=("EVERY required item, every drill, every review, every survey — with the date it is due. "
                       "Sorted by due date. Filter Status = OVERDUE to see what is behind. Fill in the last four "
                       "columns as you go; those columns are your proof at survey."))
    r = ck._hdr_row + 1
    for row in cal_rows:
        ck.append([]) if False else None
        vals = [row["due"], row["status"], row["days"], dom_names.get(row["domain"], row["domain"]),
                row["area"], row["item"], row["instance"], row["period"], row["owner"],
                row["evidence"], "", "", "", row["note"]]
        for ci, v in enumerate(vals, 1):
            c = ck.cell(r, ci, v)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = Font(size=8.5)
            c.border = THIN
            if ci == 1 and v:
                c.number_format = "yyyy-mm-dd"
        fill = (RED if row["status"] == "OVERDUE" else
                AMBER if row["status"] == "Due within 30 days" else None)
        if fill:
            for ci in range(1, 15):
                ck.cell(r, ci).fill = fill
        r += 1
    _yesno(ck, "K", ck._hdr_row + 1, r)

    # ---------- 02 calendar by month ----------
    cal = _sheet(wb, "02 Compliance Calendar",
                 ["Month", "Due date", "Item", "Which one / who", "Owner", "Section", "Done?", "Date completed"],
                 [11, 11, 56, 24, 20, 26, 9, 13],
                 notes="The same obligations arranged by month, so you can see what a given month demands.")
    r = cal._hdr_row + 1
    for row in sorted([x for x in cal_rows if x["due"]], key=lambda x: x["due"]):
        vals = [row["due"].strftime("%Y-%m"), row["due"], row["item"], row["instance"],
                row["owner"], dom_names.get(row["domain"], row["domain"]), "", ""]
        for ci, v in enumerate(vals, 1):
            c = cal.cell(r, ci, v)
            c.font = Font(size=8.5)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if ci == 2:
                c.number_format = "yyyy-mm-dd"
        r += 1
    _yesno(cal, "G", cal._hdr_row + 1, r)

    # ---------- logs ----------
    logs = [
        ("03 Drills", ["Date", "Time", "Drill type", "Site", "Shift", "Announced or unannounced",
                       "# participants", "Response time (min:sec)", "What went wrong",
                       "Corrective action", "Owner", "Date corrected", "Verified at next drill (Y/N)"],
         [11, 8, 24, 16, 14, 20, 10, 14, 36, 36, 16, 12, 14],
         "Every drill, as you run it. Fire and evacuation quarterly, on EVERY shift that delivers service, "
         "and at least one a year unannounced. A drill with no 'what went wrong' looks unexamined."),

        ("04 Incidents", ["Incident #", "Date", "Time", "Type", "Person served (initials)", "Staff",
                          "Location", "Setting", "Injury?", "Emergency services called?",
                          "Reported to supervisor (date)", "External report required?", "External report made (date)",
                          "Reviewed within 5 business days (date)", "Crisis plan updated?",
                          "Debrief with person (date)", "Debrief with staff (date)", "Corrective action", "Closed (date)"],
         [10, 11, 8, 26, 14, 16, 18, 14, 9, 12, 14, 13, 14, 15, 13, 14, 14, 34, 11],
         "Every reportable incident. The columns after 'Type' are what a surveyor traces: was it reported, "
         "reviewed, externally reported on time, debriefed, and did anything change as a result."),

        ("05 Grievances", ["Grievance #", "Date received", "Anonymous?", "Filed by (role)", "Formal/informal",
                           "Category", "Assigned to", "Acknowledged (date)", "Within 3 business days?",
                           "Decision issued (date)", "Within 30 days?", "Outcome", "Appealed?",
                           "Appeal decision (date)", "Systemic action", "Date closed"],
         [10, 12, 11, 16, 13, 20, 16, 13, 13, 13, 11, 26, 10, 14, 30, 11],
         "Log EVERY concern, including the ones resolved in two minutes at the front desk. "
         "An empty grievance log is a finding, not a clean record."),

        ("06 Record Reviews", ["Quarter", "Review date", "Reviewer", "Record #", "Person (initials)",
                               "Note writer", "Open or closed", "Items scored", "Items met", "Score %",
                               "Goal linkage OK?", "Notes timely?", "Person's own words?", "Cloned text?",
                               "Billing matches documentation?", "Findings", "Correction due",
                               "Date corrected", "Verified by", "Date verified"],
         [9, 11, 16, 10, 13, 16, 12, 11, 10, 9, 12, 11, 13, 11, 15, 34, 12, 12, 14, 12],
         "One row per record reviewed. Minimum 10% of active records or 5, whichever is greater, every quarter, "
         "with every direct service staff member represented."),

        ("07 Access and Referrals", ["Referral #", "Date received", "Source", "First contact attempted",
                                     "First contact made", "Days to first contact", "Screened (date)",
                                     "Eligible?", "First service (date)", "Days referral to service",
                                     "Engaged (2+ contacts)?", "Waitlisted?", "Days waited",
                                     "Not accepted — reason", "Alternatives given", "Warm hand-off?"],
         [10, 12, 20, 14, 14, 12, 11, 10, 13, 14, 13, 11, 11, 26, 26, 12],
         "This sheet produces your access measures (A1-A4) and your referral-out evidence in one place."),

        ("08 Satisfaction — persons served", ["Survey ID", "Date", "Anonymous?", "Q1 Respect", "Q2 Own goals",
                                              "Q3 Listened", "Q4 Rights explained", "Q5 Knows how to complain",
                                              "Q6 Easy to start", "Q7 Where/when works", "Q8 Culture respected",
                                              "Q9 Pathway respected", "Q10 More people to lean on",
                                              "Q11 Closer to life I want", "Q12 Would recommend",
                                              "What helped most", "What to change"],
         [9, 11, 11] + [11] * 12 + [34, 34],
         "Enter 1-5 for each question (or blank for N/A). The analysis averages these and charts them. "
         "Also record how many surveys you ISSUED on the '15 Measure Results' sheet so response rate is honest."),

        ("09 Satisfaction — stakeholders", ["Survey ID", "Date", "Stakeholder type", "Q1 Easy to refer",
                                            "Q2 Contacted quickly", "Q3 Communication back", "Q4 Does what it says",
                                            "Q5 Staff prepared", "Q6 Paperwork meets requirements",
                                            "Q7 Problems resolved", "Q8 People better off", "Q9 Would refer again",
                                            "What works", "What should change"],
         [9, 11, 22] + [12] * 9 + [34, 34],
         "Referral sources, payers, community partners, families, staff. Stakeholder input is a separate "
         "requirement from person-served satisfaction — you need both."),

        ("10 Outcomes — persons served", ["Record #", "Initials", "Admission date", "Discharge date",
                                          "Reason for discharge", "Goals set", "Goals achieved",
                                          "Housing goal? Achieved?", "Employment/education goal? Achieved?",
                                          "Benefits goal? Achieved?", "Natural supports at admission",
                                          "Natural supports at last review", "Natural supports at discharge",
                                          "Days to first plan", "Plan reviews due", "Plan reviews on time",
                                          "Follow-up attempted (date)", "Follow-up reached?", "Status at follow-up"],
         [10, 10, 12, 12, 24, 9, 11, 16, 20, 15, 16, 17, 16, 12, 12, 14, 15, 12, 24],
         "One row per person served. This single sheet produces almost every effectiveness measure "
         "CARF asks about, including whether people's unpaid support networks actually grew."),

        ("11 Personnel", ["Name", "Role", "Hire date", "Credential", "Credential #", "Credential expires",
                          "Background check date", "Exclusion screen (last)", "Orientation completed",
                          "Competency — initial", "Competency — last annual", "Direct observation (last)",
                          "Supervision — last", "Supervision count last 12 mo", "Appraisal — last",
                          "Disclosure plan (date)", "Wellness plan (date)", "First aid/CPR expires",
                          "Annual training complete?", "Separation date"],
         [18, 20, 11, 24, 13, 14, 14, 14, 14, 14, 15, 14, 13, 15, 12, 14, 14, 15, 15, 12],
         "One row per employee, contractor, student, intern and volunteer with contact. This is the sheet "
         "that tells you, before the surveyor does, whose file is short."),

        ("12 Financial", ["Month", "Budget revenue", "Actual revenue", "Budget expense", "Actual expense",
                          "Variance %", "Cash on hand", "Days cash", "AR over 90 days",
                          "Claims submitted", "Claims denied", "Denial rate %", "Overpayments identified",
                          "Overpayments repaid", "Bank reconciled by", "Date reconciled", "Reviewed by"],
         [10, 14, 14, 14, 14, 11, 13, 10, 14, 13, 12, 12, 15, 15, 18, 13, 16],
         "Monthly. Produces business-function measures B1 and B2, and proves the separation-of-duties control."),

        ("13 Accessibility", ["Date identified", "Area (of the 8)", "Barrier", "How it was identified",
                              "Action", "Owner", "Target date", "Status", "Date removed/reduced",
                              "Cost", "In budget?", "If not removed — interim accommodation"],
         [13, 20, 36, 22, 36, 18, 12, 14, 15, 10, 11, 34],
         "All eight areas: architectural, environmental, attitudinal, financial, employment, communication, "
         "transportation, digital. An area with no row looks unexamined — enter 'None identified' and say how you looked."),

        ("14 PI Projects", ["Project #", "Started", "Problem (with a number)", "Source of the problem",
                            "Aim (from X to Y by when)", "Change tested", "Owner", "Baseline", "Mid-point",
                            "End result", "Decision (adopt/adapt/abandon)", "What changed permanently",
                            "Date embedded", "Verified still in place"],
         [10, 11, 40, 22, 34, 34, 16, 11, 11, 11, 20, 34, 13, 15],
         "At least one completed project with before-and-after numbers before survey. This is where "
         "'we improve' stops being a claim and becomes evidence."),
    ]
    for title, headers, widths, note in logs:
        _sheet(wb, title, headers, widths, notes=note)

    # ---------- 15 measure results ----------
    mr = _sheet(wb, "15 Measure Results",
                ["Measure ID", "Measure", "Domain", "Target", "Unit", "Q1", "Q2", "Q3", "Q4",
                 "Year", "Prior year", "Source", "Owner"],
                [10, 52, 20, 10, 12, 9, 9, 9, 9, 9, 11, 24, 20],
                notes=("Enter the ACTUAL number for each quarter. Leave a cell blank if you genuinely do not "
                       "have it yet — analyze.py will say so rather than guess. These cells drive every chart "
                       "and the annual analysis."))
    measures = [
        ("E1", "Persons achieving at least one plan goal", "Effectiveness", "", "%"),
        ("E2", "Persons whose unpaid natural supports increased", "Effectiveness", "", "%"),
        ("E3", "Persons obtaining or maintaining housing (of those with a housing goal)", "Effectiveness", "", "%"),
        ("E4", "Persons obtaining employment or education (of those with that goal)", "Effectiveness", "", "%"),
        ("E5", "Persons reporting improved quality of life", "Effectiveness", "", "%"),
        ("F1", "Average length of service to first goal achievement", "Efficiency", "", "days"),
        ("F2", "Scheduled contacts kept", "Efficiency", "", "%"),
        ("F3", "Active persons served per direct service FTE", "Efficiency", "", "count"),
        ("F4", "Cost per person served", "Efficiency", "", "$"),
        ("A1", "Mean days from referral to first contact", "Access", "", "days"),
        ("A2", "Mean days from first contact to signed plan", "Access", "", "days"),
        ("A3", "Referrals engaged (2+ contacts)", "Access", "", "%"),
        ("A4", "Mean days waited on the waiting list", "Access", "", "days"),
        ("S1", "Overall satisfaction of persons served", "Satisfaction", "", "%"),
        ("S2", "Would recommend the service", "Satisfaction", "", "%"),
        ("S3", "Survey response rate — persons served", "Satisfaction", "", "%"),
        ("S4", "Stakeholder satisfaction", "Satisfaction", "", "%"),
        ("B1", "Budget to actual variance", "Business", "", "%"),
        ("B2", "Claim denial rate", "Business", "", "%"),
        ("B3", "Notes completed within the required timeframe", "Business", "", "%"),
        ("B4", "Staff turnover", "Business", "", "%"),
        ("B5", "Staff with all required training current", "Business", "100", "%"),
        ("B6", "Staff with current credential", "Business", "100", "%"),
        ("B7", "Critical incidents per 100 service contacts", "Business", "", "rate"),
        ("B8", "Grievances resolved within 30 days", "Business", "100", "%"),
        ("B9", "Record review findings corrected and verified", "Business", "100", "%"),
        ("B10", "Drills completed on schedule, all shifts", "Business", "100", "%"),
    ]
    r = mr._hdr_row + 1
    for mid, name, dom, target, unit in measures:
        for ci, v in enumerate([mid, name, dom, target, unit], 1):
            c = mr.cell(r, ci, v)
            c.font = Font(size=9, bold=(ci == 1))
            c.alignment = Alignment(wrap_text=True, vertical="top")
        for ci in range(6, 12):
            mr.cell(r, ci).fill = SUB
            mr.cell(r, ci).border = THIN
        r += 1

    # ---------- 16 evidence register ----------
    ev = _sheet(wb, "16 Evidence Register",
                ["#", "Document / evidence", "CARF area", "Section (domain)", "Type",
                 "Do we have it?", "Date on the document", "Covers period", "Expires / next due",
                 "Where it is filed", "Binder tab", "Owner", "Gap or what is missing", "Action", "Status"],
                [5, 52, 9, 26, 18, 12, 14, 16, 15, 30, 10, 20, 30, 30, 12],
                notes=("Every document a surveyor may ask for, and whether you actually hold it. "
                       "Fill 'Do we have it?', the date, and where it lives. This becomes your survey-day index."))
    return wb, ev, mr


def fill_evidence_register(ev, rows):
    """rows: (label, area, domain, type, owner) tuples from the content set."""
    r = ev._hdr_row + 1
    for i, (label, area, domain, typ, owner) in enumerate(rows, 1):
        vals = [i, label, area, domain, typ, "", "", "", "", "", area, owner, "", "", "Not started"]
        for ci, v in enumerate(vals, 1):
            c = ev.cell(r, ci, v)
            c.font = Font(size=8.5)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = THIN
        r += 1
    _yesno(ev, "F", ev._hdr_row + 1, r)
    return r
