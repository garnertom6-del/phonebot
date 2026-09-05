#!/usr/bin/env python3
"""Build one provider's complete CARF accreditation packet.

    python3 _engine/build_provider.py <slug> [--no-pdf]

Everything provider-specific comes from providers/<slug>/provider.json.
Nothing in _engine/ is ever edited per provider.
"""
import csv
import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import resolve as R
import docx_kit as K
import pdf_kit as P
import json

SKILL = R.SKILL

UNVERIFIED = ("DRAFT - the CARF section map behind this document has NOT been checked "
              "against the purchased manual. Set manual_verified: true before use.")

AREA_ORDER = ["1.A", "1.B", "1.C", "1.D", "1.E", "1.F", "1.G", "1.H", "1.I", "1.J",
              "1.K", "1.L", "1.M", "1.N", "2.A", "2.B", "2.C", "2.D", "2.E", "2.F",
              "2.G", "2.H", "3.CI", "3.PEER"]

PLAN_ORDER = ["strategic", "risk", "safety", "continuity", "technology", "accessibility",
              "cultural", "compliance", "financial", "workforce",
              "performance-measurement", "performance-improvement"]


class Sink:
    """One writing surface, two back ends.

    Each document is built once against this interface and emitted as both DOCX
    (editable, for the agency) and PDF (print-ready, for the board and the
    surveyor). Generating the PDF natively means the packet does not depend on
    LibreOffice or Word being installed anywhere.
    """

    def __init__(self, kind):
        self.kind = kind
        if kind == "docx":
            self.doc = K.new_document()
        else:
            self.story = []

    def title_page(self, *a, **kw):
        (K.title_page(self.doc, *a, **kw) if self.kind == "docx"
         else P.title_page(self.story, *a, **kw))

    def heading(self, text, level=1):
        (K.heading(self.doc, text, level) if self.kind == "docx"
         else P.heading(self.story, text, level))

    def md(self, text, base_level=1):
        (K.render_markdown(self.doc, text, base_level) if self.kind == "docx"
         else P.render_markdown(self.story, text, base_level))

    def table(self, rows):
        (K._table(self.doc, rows) if self.kind == "docx" else P.table(self.story, rows))

    def notice(self, text):
        (K.notice(self.doc, text) if self.kind == "docx" else P.notice(self.story, text))

    def image(self, path, width_in=6.6):
        (K.image(self.doc, path, width_in) if self.kind == "docx"
         else P.image(self.story, path, width_in))

    def pb(self):
        (K.page_break(self.doc) if self.kind == "docx" else P.page_break(self.story))

    def rule(self):
        (K.rule(self.doc) if self.kind == "docx" else P.rule(self.story))

    def space(self):
        (self.doc.add_paragraph() if self.kind == "docx"
         else self.story.append(P.Spacer(1, 6)))

    def save(self, path, footer):
        if self.kind == "docx":
            K.footer_text(self.doc, footer)
            self.doc.save(path)
        else:
            P.build(path, self.story, footer)


def sub(text, tk):
    return R.substitute(text, tk)


def load(kind, name, tk):
    """Read a content file with tokens already substituted, then split front matter.

    Substituting BEFORE the split matters: front-matter values such as
    `when: Within {{PLAN_DAYS}} days` are rendered into headers and contents tables.
    """
    return R.parse_front_matter(sub(R.read_content(kind, name), tk))


# --------------------------------------------------------------------------
def build_manual(s, tk, smap, verified):
    s.title_page(tk["AGENCY"], "Policy and Procedure Manual",
        f"Organized to the CARF Behavioral Health Standards Manual, {tk['MANUAL_YEAR']} edition",
        [f"Program: {tk['PROGRAM_NAME']}",
         f"Effective date: {tk['EFFECTIVE_DATE']}",
         f"Approved by: {tk['GOVERNING_BODY']}",
         f"Generated {tk['TODAY']}" + (f" by {tk['CONSULTANT']}" if tk["CONSULTANT"] else "")],
        warning=None if verified else UNVERIFIED)

    s.heading("How to use this manual", 1)
    for line in [
        "This manual is organized by CARF standard area so that a surveyor asking about an area "
        "can be handed the matching tab. Each policy names its owner, its review frequency, and "
        "-- most importantly -- **the evidence a surveyor will ask for**. A policy with no "
        "evidence behind it is the single most common reason an organization does not conform.",
        "**Anything in [SQUARE BRACKETS] or on a fill-in line must be completed by the agency "
        "before this manual is adopted.** See BLANKS_TO_COMPLETE.md for the full list.",
        "**Before adoption:** read every policy. These are written to be defensible and to fit a "
        "peer-delivered community program, but they are your agency's statements about your "
        "agency. Change anything that is not true of you, and change your practice where the "
        "policy describes what you intend to do.",
    ]:
        s.md(line)
    s.pb()

    s.heading("Contents", 1)
    by_area = {}
    for pid in R.list_content("policies"):
        meta, _ = load("policies", pid, tk)
        by_area.setdefault(meta.get("area", "?"), []).append((pid, meta))
    rows = [["Area", "Policy #", "Title", "Owner", "Review"]]
    for area in AREA_ORDER:
        info = next((a for a in smap["areas"] if a["code"] == area), None)
        name = info["name"] if info else ""
        for pid, meta in sorted(by_area.get(area, [])):
            rows.append([f"{area} {name}", pid, meta.get("title", ""),
                         sub(meta.get("owner", ""), tk), meta.get("review", "")])
    s.table(rows)
    s.pb()

    for area in AREA_ORDER:
        info = next((a for a in smap["areas"] if a["code"] == area), None)
        if not info or area not in by_area:
            continue
        s.heading(f"Section {area} — {info['name']}", 1)
        s.rule()
        s.md(f"**What this area is about:** {sub(info['plain'], tk)}")
        s.md("**Questions a surveyor asks in this area:**")
        for q in info["surveyor_asks"]:
            s.md(f"- {q}")
        s.md(f"**Who gets interviewed:** {', '.join(info['interviews'])}")
        if info.get("note"):
            s.notice(info["note"])
        s.space()

        for pid, meta in sorted(by_area[area]):
            _, body = load("policies", pid, tk)
            s.heading(f"{pid} — {sub(meta.get('title',''), tk)}", 2)
            s.table([["Policy #", "Owner", "Review", "Standard area"],
                           [pid, sub(meta.get("owner", ""), tk),
                            meta.get("review", ""), f"{area} {info['name']}"]])
            s.md(sub(body, tk), base_level=2)
            s.md(
                              "Approved: ____________________________   Date: ______________   "
                              "Next review: ______________")
            s.pb()


# --------------------------------------------------------------------------
def build_plans(s, tk, verified):
    s.title_page(tk["AGENCY"], "Annual Written Plans",
        "The twelve plans CARF expects to see written, implemented, and reviewed",
        [f"Program: {tk['PROGRAM_NAME']}", f"Generated {tk['TODAY']}"],
        warning=None if verified else UNVERIFIED)

    s.heading("Why these twelve", 1)
    s.md(
        "CARF does not just ask whether you have a policy. For these areas it asks for a **written "
        "plan** -- a document with objectives, named owners, dates, and evidence that it was "
        "reviewed and acted on. A plan dated more than twelve months ago is treated as no plan. "
        "A plan with no documented review is treated as no plan.")
    s.md(
        "**These are templates.** Everything in [SQUARE BRACKETS] is yours to fill in. A template "
        "handed to a surveyor with brackets still in it is worse than no plan at all.")
    s.pb()

    rows = [["#", "Plan", "CARF area", "Cycle", "Approved by"]]
    metas = {}
    for i, pid in enumerate(PLAN_ORDER, 1):
        meta, _ = load("plans", pid, tk)
        metas[pid] = meta
        rows.append([str(i), meta.get("title", pid), meta.get("area", ""),
                     meta.get("cycle", ""), sub(meta.get("approver", ""), tk)])
    s.heading("Contents", 1)
    s.table(rows)
    s.pb()

    for pid in PLAN_ORDER:
        meta = metas[pid]
        _, body = load("plans", pid, tk)
        s.heading(sub(meta.get("title", pid), tk), 1)
        s.table([["CARF area", "Review cycle", "Approved by"],
                       [meta.get("area", ""), meta.get("cycle", ""),
                        sub(meta.get("approver", ""), tk)]])
        s.md(sub(body, tk), base_level=2)
        s.pb()


# --------------------------------------------------------------------------
def build_forms(s, tk, smap, verified):
    s.title_page(tk["AGENCY"], "Forms and Tools Packet",
        "Every form the policies and plans refer to",
        [f"Program: {tk['PROGRAM_NAME']}", f"Generated {tk['TODAY']}"],
        warning=None if verified else UNVERIFIED)

    area_of_form = {}
    for a in smap["areas"]:
        for f in a["forms"]:
            area_of_form.setdefault(f, a["code"])

    names = R.list_content("forms")
    metas = {}
    rows = [["#", "Form", "CARF area", "Who completes it", "When", "Where it is filed"]]
    for i, fid in enumerate(names, 1):
        meta, _ = load("forms", fid, tk)
        metas[fid] = meta
        rows.append([str(i), sub(meta.get("title", fid), tk),
                     meta.get("area", area_of_form.get(fid, "")),
                     sub(meta.get("who", ""), tk), meta.get("when", ""),
                     sub(meta.get("file", ""), tk)])
    s.heading("Contents", 1)
    s.md(
        "**Print what you use. Use what you print.** A form in this packet that your staff do not "
        "actually complete is a finding waiting to happen -- either start using it or take it out "
        "and remove the reference from the policy.")
    s.table(rows)
    s.pb()

    for fid in names:
        meta = metas[fid]
        _, body = load("forms", fid, tk)
        s.heading(sub(meta.get("title", fid), tk), 1)
        s.table([["CARF area", "Who completes it", "When", "Filed in"],
                       [meta.get("area", ""), sub(meta.get("who", ""), tk),
                        meta.get("when", ""), sub(meta.get("file", ""), tk)]])
        s.md(sub(body, tk), base_level=2)
        s.pb()


# --------------------------------------------------------------------------
def build_roadmap(s, tk, smap, verified):
    s.title_page(tk["AGENCY"], "Roadmap, Training Matrix and Survey Preparation",
        "What to do, in what order, and how to rehearse for the survey",
        [f"Target survey: {tk['SURVEY_TARGET']}", f"Generated {tk['TODAY']}"],
        warning=None if verified else UNVERIFIED)

    for mid in ["timeline", "training-matrix", "survey-prep"]:
        _, body = load("meta", mid, tk)
        s.md(sub(body, tk), base_level=1)
        s.pb()

    s.heading("Evidence Binder Index", 1)
    s.md(
        "One tab per standard area. On survey day the surveyor names an area; you hand them that "
        "tab. Build this as you go, not in the last month.")
    for a in smap["areas"]:
        s.heading(f"Tab {a['code']} — {a['name']}", 2)
        s.md(sub(a["plain"], tk))
        rows = [["What goes in this tab", "Type", "Present?", "Date on it"]]
        for pid in a["policies"]:
            meta, _ = load("policies", pid, tk)
            rows.append([f"{pid} {sub(meta.get('title',''), tk)}", "Policy", "", ""])
        for pl in a["plans"]:
            meta, _ = load("plans", pl, tk)
            rows.append([sub(meta.get("title", pl), tk), "Plan", "", ""])
        for f in a["forms"]:
            meta, _ = load("forms", f, tk)
            rows.append([sub(meta.get("title", f), tk), "Completed form / evidence", "", ""])
        s.table(rows)
        s.md("**Surveyor will ask:** " + "  ".join(f"({i+1}) {q}"
                          for i, q in enumerate(a["surveyor_asks"])))
        s.space()


# --------------------------------------------------------------------------
def build_interviews(s, tk, verified):
    s.title_page(
        tk["AGENCY"], "Surveyor Interview Bank",
        "What the surveyor asks, what a strong answer sounds like, and what sinks you",
        [f"Program: {tk['PROGRAM_NAME']}", f"Generated {tk['TODAY']}"],
        warning=None if verified else UNVERIFIED)
    _, body = load("meta", "interview-bank", tk)
    s.md(body, base_level=1)


# --------------------------------------------------------------------------
def build_checklist(path, tk, smap):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Self-Study"
    head = ["CARF area", "Area name", "What this area requires (plain English)",
            "Required document / evidence", "Type", "Do we have it?",
            "Date on the document", "Owner", "Gap / what is missing",
            "Action needed", "Due date", "Status"]
    ws.append(head)
    navy = PatternFill("solid", fgColor="1F3B63")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = navy
        c.alignment = Alignment(wrap_text=True, vertical="top")

    for a in smap["areas"]:
        items = []
        for pid in a["policies"]:
            m, _ = load("policies", pid, tk)
            items.append((f"{pid} {sub(m.get('title',''), tk)}", "Policy", sub(m.get("owner", ""), tk)))
        for pl in a["plans"]:
            m, _ = load("plans", pl, tk)
            items.append((sub(m.get("title", pl), tk), "Written plan", sub(m.get("approver", ""), tk)))
        for f in a["forms"]:
            m, _ = load("forms", f, tk)
            items.append((sub(m.get("title", f), tk), "Form / evidence", sub(m.get("who", ""), tk)))
        for label, typ, owner in items:
            ws.append([a["code"], a["name"], sub(a["plain"], tk), label, typ,
                       "", "", owner, "", "", "", "Not started"])

    widths = [9, 26, 55, 46, 18, 14, 16, 24, 30, 30, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Interview Prep")
    ws2.append(["CARF area", "Who gets interviewed", "Question the surveyor asks",
                "Our answer (write it in their own words)", "Ready?"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = navy
        c.alignment = Alignment(wrap_text=True, vertical="top")
    for a in smap["areas"]:
        for q in a["surveyor_asks"]:
            ws2.append([a["code"], ", ".join(a["interviews"]), q, "", ""])
    for i, w in enumerate([9, 30, 70, 60, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for row in ws2.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.freeze_panes = "A2"

    wb.save(path)


# --------------------------------------------------------------------------
def evidence_rows(tk, smap):
    """Every document a surveyor may ask for, as rows for the Evidence Register."""
    dom = {d["key"]: d["name"] for d in json.load(
        open(os.path.join(R.CONTENT, "meta", "obligations.json")))["domains"]}
    area_dom = {}
    for d in json.load(open(os.path.join(R.CONTENT, "meta", "obligations.json")))["domains"]:
        for a in d["areas"]:
            area_dom[a] = d["name"]
    rows = []
    for a in smap["areas"]:
        dname = area_dom.get(a["code"], a["name"])
        for pid in a["policies"]:
            m, _ = load("policies", pid, tk)
            rows.append((f"{pid} {m.get('title','')}", a["code"], dname, "Policy",
                         m.get("owner", "")))
        for pl in a["plans"]:
            m, _ = load("plans", pl, tk)
            rows.append((m.get("title", pl), a["code"], dname, "Written plan",
                         m.get("approver", "")))
        for f in a["forms"]:
            m, _ = load("forms", f, tk)
            rows.append((f"COMPLETED: {m.get('title', f)}", a["code"], dname,
                         "Completed evidence", m.get("who", "")))
    return rows


def write_blanks(path, res, tk, verified, provider):
    lines = [f"# BLANKS TO COMPLETE — {tk['AGENCY']}", "",
             f"Generated {tk['TODAY']}.  **Read this first.**", "",
             "This packet never invents a name, a date, a credential, or a number. "
             "Everything below is either missing from the provider record or is a choice "
             "the agency has to confirm. Nothing here is optional.", ""]

    lines += ["## 0. The verification gate", ""]
    if verified:
        lines += ["- [x] The CARF section map has been checked against the purchased "
                  f"{provider.get('manual_year')} manual (`manual_verified: true`).", ""]
    else:
        lines += [
            "- [ ] **THE SECTION MAP HAS NOT BEEN VERIFIED.** Every document in this packet is "
            "stamped DRAFT until it is.",
            "",
            "  This skill organizes your documents by the CARF Behavioral Health Standards "
            "Manual's long-standing outline. It does **not** contain CARF's copyrighted standard "
            "text, and nobody here has read your edition. Do this:",
            "",
            f"  1. Buy the {provider.get('manual_year')} Behavioral Health Standards Manual from carf.org.",
            "  2. Open `_engine/content/meta/standards_map.json`.",
            "  3. Check each area letter and name against the manual's contents pages. Fix any "
            "that moved, merged, split, or were renamed.",
            "  4. Read each area's standards and add anything this skill does not cover to the "
            "self-study checklist.",
            "  5. Set `manual_verified: true` in provider.json and rebuild.", ""]

    if res.blanks:
        lines += ["## 1. Missing from the provider record", "",
                  "Each of these prints as a blank line in the documents. Add it to "
                  "`provider.json` and rebuild — do not hand-edit the Word files, or the next "
                  "build will overwrite you.", "",
                  "| Field in provider.json | What it is | Why it matters |", "|---|---|---|"]
        for k, label, matters in res.blanks:
            lines.append(f"| `{k}` | {label} | {matters} |")
        lines.append("")
    else:
        lines += ["## 1. Missing from the provider record", "",
                  "Nothing missing. Every field was supplied.", ""]

    if res.decisions:
        lines += ["## 2. Decisions we defaulted — confirm each one", "",
                  "These are **your agency's operational choices**, not CARF numbers. The default "
                  "is a reasonable starting point. Your state licensure rule or a payer contract "
                  "may be stricter, and the strictest one wins. Once confirmed, put them in a "
                  "`timeframes` object in provider.json and rebuild.", "",
                  "| Token | Default used | What to check |", "|---|---|---|"]
        for k, v, why in res.decisions:
            lines.append(f"| `{k}` | {v} | {why} |")
        lines.append("")

    lines += ["## 3. Bracketed blanks inside the documents", "",
              "The annual plans and several forms carry `[SQUARE BRACKET]` prompts that only the "
              "agency can answer — budget figures, insurance carriers, drill dates, demographic "
              "percentages, caseload rationale. Work through them plan by plan. **A plan handed to "
              "a surveyor with brackets still in it is worse than no plan.**", "",
              "## 4. Signatures", "",
              "Every policy carries an approval line and every plan a signature line. They are "
              "left blank deliberately. Nobody signs on the agency's behalf but the agency.", "",
              "## 5. Read before adopting", "",
              "These documents make flat factual claims about your agency — that it does not "
              "administer medication, does not use restraint, does or does not transport people. "
              "Those sentences are true only because your provider record says so. Read them. If "
              "one is not true of you, fix the record and rebuild; never edit the sentence out.", ""]
    with open(path, "w") as f:
        f.write("\n".join(lines))


def write_start_here(path, tk, verified, outdir, n_items=0, n_overdue=0):
    md = f"""# START HERE — {tk['AGENCY']}
### Your CARF accreditation system, and exactly what to do with it

Generated {tk['TODAY']}. Target survey: {tk['SURVEY_TARGET']}.

---

## First, the two things you must know

**1. Nobody here has read your CARF manual.** CARF's standards are copyrighted and sold by
CARF. This packet is organized to the manual's long-standing structure and is written from
scratch — it is not a copy of the standards and it is not a substitute for buying the manual.
Buy the {tk['MANUAL_YEAR']} Behavioral Health Standards Manual from carf.org, check the section
map, then set `manual_verified: true` and rebuild. Until you do, every document says DRAFT.

**2. CARF accredits PROGRAMS, not job titles.** There is no "peer support accreditation."
You apply for a program — for you, **{tk['PROGRAM_NAME']}** — and peer support is the workforce
and the service model that delivers it. The peer-specific requirements live in Section 1.I
(credential, competency, supervision), Section 2.A and 2.C (program description, planning with
the person), and the Section 3 program standards. This packet covers all of them, and adds a
cross-cutting peer section so nothing peer-specific falls between the cracks.

---

## What is in this folder

| File | What it is | What to do with it |
|---|---|---|
| `BLANKS_TO_COMPLETE.md` | Everything missing or defaulted | **Read this before anything else** |
| `01_Policy_and_Procedure_Manual` | 58 policies, organized by CARF area | Read, edit to fit you, adopt, sign |
| `02_Annual_Plans` | The 12 written plans CARF expects | Fill in every [BRACKET], sign, date |
| `03_Forms_Packet` | 72 forms the policies refer to | Print the ones you will actually use |
| `04_Roadmap_and_Survey_Prep` | 12-month countdown, training matrix, evidence binder index | Work the roadmap |
| `05_Self_Study_Checklist.xlsx` | Every required document, with a status column | Your conformance tracker |
| `06_Surveyor_Interview_Bank` | What surveyors ask staff and the people you serve, with the shape of a strong answer and what sinks you | Rehearse with it 60 days out |
| `07_Performance_Analysis_Report` | Built from YOUR data, with charts | Re-run it monthly |

And one level up, in **`../data/`**:

| File | What it is |
|---|---|
| `Evidence_and_Data_Workbook.xlsx` | **The spine of the whole thing.** {n_items} dated required items, a compliance calendar, 12 data logs, the measure set, and the evidence register |

---

## Do it in this order

**Step 1 — Buy the manual.** carf.org. You cannot do this properly without it.

**Step 2 — Open `BLANKS_TO_COMPLETE.md`.** Fill in what is missing in
`providers/{tk.get('SLUG','<slug>')}/provider.json`, then rebuild:
`python3 _engine/build_provider.py {tk.get('SLUG','<slug>')}`. Do not hand-edit the Word files
first — a rebuild overwrites them. (Your workbook is never overwritten.)

**Step 3 — Verify the section map** against the manual. Set `manual_verified: true`. Rebuild.
The DRAFT stamp disappears.

**Step 4 — Read the whole policy manual.** Every policy is a statement about your agency.
Change anything that is not true of you. Where it describes what you *intend* to do, change
your practice to match — or change the policy.

**Step 5 — Adopt it.** Your governing body approves the manual in a meeting, and it goes in
the minutes. Date every policy.

**Step 6 — Open the workbook and work sheet `01 MASTER CHECK-OFF LIST`.**
It has **{n_items} dated items** — every drill, every review, every survey, every plan, every
report — with the date each is due and who owns it.{'' if not n_overdue else f" **{n_overdue} are already overdue.**"}
Filter Status = OVERDUE and start there.

**Step 7 — Start generating evidence today.** This is the part people get wrong. CARF does not
want to see that you wrote a policy last month. It wants to see that you have been *running*
this — drills, record reviews, supervision notes, incident analyses, satisfaction results, a
completed annual analysis. Enter each one in the workbook the day it happens.

**Step 8 — Run the analysis monthly.**
`python3 _engine/analyze.py {tk.get('SLUG','<slug>')}`
That reads your workbook and produces `07_Performance_Analysis_Report` with real charts,
real trends, and an honest list of what is still missing. It never invents a number.

**Step 9 — Mock survey, at least 60 days out.** Use `06_Surveyor_Interview_Bank`. Have someone
outside the program ask the questions. Write down the real answers. Every shrug is a gap —
and you have found it before the surveyor did.

---

## The five things that most often go wrong

1. **Goals written in staff language.** A surveyor reads a goal aloud and asks the person
   served if those are their words. Write goals in quotation marks, in the person's voice.
2. **Drills on one shift only.** Every shift that delivers service needs drills, and at least
   one a year must be unannounced.
3. **Results never shared.** Analysing performance is half of area 1.N. Telling staff and the
   people you serve what you found is the other half, and it is the most-missed requirement
   in the manual.
4. **Training with no competency.** A sign-in sheet proves attendance. CARF asks how you know
   the person can do the work. Direct observation, documented.
5. **Findings that were never closed.** Finding a problem is half. The tracker must show
   somebody verified the fix.

---

## The one rule that matters more than all of this

**Never write a date for something that did not happen, and never enter a number you cannot
trace to a source.** A gap you can explain is survivable at survey. A fabricated record is
fraud, and it is the one thing that ends an agency. Every tool here is built to make a gap
visible rather than paper over it — that is the point of it.
"""
    with open(path, "w") as f:
        f.write(md)


def update_tracker(provider, tk, verified, nblanks):
    path = os.path.join(SKILL, "TRACKER.csv")
    head = ["slug", "legal_name", "programs", "manual_year", "manual_verified",
            "accreditation_type", "target_survey_month", "six_months_of_data_start",
            "open_blanks", "last_built"]
    rows = []
    if os.path.exists(path):
        with open(path) as f:
            rows = [r for r in csv.DictReader(f) if r.get("slug") != provider["slug"]]
    rows.append({
        "slug": provider["slug"], "legal_name": provider["legal_name"],
        "programs": "|".join(provider.get("programs", [])),
        "manual_year": provider.get("manual_year", ""),
        "manual_verified": str(verified),
        "accreditation_type": provider.get("accreditation_type") or "",
        "target_survey_month": provider.get("target_survey_month") or "",
        "six_months_of_data_start": provider.get("six_months_of_data_start") or "",
        "open_blanks": str(nblanks), "last_built": date.today().isoformat()})
    rows.sort(key=lambda r: (r.get("target_survey_month") or "9999-99", r["slug"]))
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=head)
        w.writeheader()
        w.writerows(rows)
    return path



# --------------------------------------------------------------------------
def main():
    if len(sys.argv) < 2:
        sys.exit("usage: build_provider.py <slug> [--no-pdf]")
    slug = sys.argv[1]
    want_pdf = "--no-pdf" not in sys.argv  # kept for compatibility; PDFs are native now

    provider = R.load_provider(slug)
    provider.setdefault("slug", slug)
    res = R.Resolver(provider)
    try:
        res.gate()
    except R.GateError as e:
        print("\nCAPABILITY GATE — build refused.\n")
        print(str(e))
        print("\nThe generated manual makes flat factual claims about the agency. Building a\n"
              "provider against content that contradicts its real profile files a document that\n"
              "says something untrue. Write the missing content first.\n"
              "NEVER edit the provider record to make the gate pass.\n")
        sys.exit(2)

    tk = res.tokens()
    tk["SLUG"] = slug
    verified = bool(provider.get("manual_verified"))
    smap = json.load(open(os.path.join(R.CONTENT, "meta", "standards_map.json")))

    outdir = os.path.join(SKILL, "providers", slug, "output")
    os.makedirs(outdir, exist_ok=True)
    foot = f"{tk['AGENCY_SHORT']} — CARF packet — {'' if verified else 'DRAFT, section map unverified — '}generated {tk['TODAY']}"

    docs = [
        ("01_Policy_and_Procedure_Manual", lambda s: build_manual(s, tk, smap, verified)),
        ("02_Annual_Plans", lambda s: build_plans(s, tk, verified)),
        ("03_Forms_Packet", lambda s: build_forms(s, tk, smap, verified)),
        ("04_Roadmap_and_Survey_Prep", lambda s: build_roadmap(s, tk, smap, verified)),
        ("06_Surveyor_Interview_Bank", lambda s: build_interviews(s, tk, verified)),
    ]
    written = []
    for name, fn in docs:
        for kind, ext in (("docx", ".docx"), ("pdf", ".pdf")):
            sink = Sink(kind)
            fn(sink)
            path = os.path.join(outdir, name + ext)
            try:
                sink.save(path, foot)
                written.append(path)
            except Exception as e:
                print(f"  ! {kind.upper()} failed for {name}: {e}")
        print(f"  wrote {name}.docx + .pdf")

    xlsx = os.path.join(outdir, "05_Self_Study_Checklist.xlsx")
    build_checklist(xlsx, tk, smap)
    print("  wrote 05_Self_Study_Checklist.xlsx")

    # --- the evidence and data workbook + the dated compliance calendar ---
    import calendar_engine as CE
    import workbook as WB
    cal_rows, cal_start, cal_end = CE.expand(provider)
    data_dir = os.path.join(SKILL, "providers", slug, "data")
    os.makedirs(data_dir, exist_ok=True)
    wb_path = os.path.join(data_dir, "Evidence_and_Data_Workbook.xlsx")
    if os.path.exists(wb_path) and "--new-workbook" not in sys.argv:
        print("  KEPT existing Evidence_and_Data_Workbook.xlsx (it holds your entered data).")
        print("       Pass --new-workbook to build a fresh one alongside it.")
    else:
        if os.path.exists(wb_path):
            stamp = date.today().isoformat()
            wb_path = os.path.join(data_dir, f"Evidence_and_Data_Workbook_{stamp}.xlsx")
        wbk, ev, _ = WB.build(wb_path, tk, provider, cal_rows, cal_start, cal_end)
        WB.fill_evidence_register(ev, evidence_rows(tk, smap))
        wbk.save(wb_path)
        print(f"  wrote {os.path.basename(wb_path)}  ({len(cal_rows)} dated items on the check-off list)")

    n_overdue = sum(1 for r in cal_rows if r["status"] == "OVERDUE")
    write_start_here(os.path.join(outdir, "00_START_HERE.md"), tk, verified, outdir,
                     len(cal_rows), n_overdue)
    write_blanks(os.path.join(outdir, "BLANKS_TO_COMPLETE.md"), res, tk, verified, provider)
    print("  wrote 00_START_HERE.md, BLANKS_TO_COMPLETE.md")


    tracker = update_tracker(provider, tk, verified, len(res.blanks))

    print(f"\nPacket built: {outdir}")
    if not verified:
        print("\n  ** manual_verified is false — every document is stamped DRAFT. **")
    print(f"  {len(res.blanks)} missing field(s), {len(res.decisions)} defaulted decision(s).")
    print(f"  Start with 00_START_HERE.md, then BLANKS_TO_COMPLETE.md.")
    print(f"  Then work providers/{slug}/data/Evidence_and_Data_Workbook.xlsx sheet 01.")
    print(f"  When it has data in it:  python3 _engine/analyze.py {slug}")
    print(f"  Tracker updated: {tracker}")


if __name__ == "__main__":
    main()
